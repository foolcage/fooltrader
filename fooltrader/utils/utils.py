import datetime
import itertools
import json
import logging
import os

import openpyxl
import pandas as pd

from fooltrader import settings
from fooltrader.contract.data_contract import TICK_COLUNM
from fooltrader.contract.files_contract import get_kdata_path, get_kdata_dir, get_kdata_path_ths, \
    get_trading_dates_path_sse, get_trading_dates_path_ths, get_trading_dates_path, get_tick_path, \
    get_tick_dir, get_sh_stock_list_path, get_sz_stock_list_path, get_finance_dir, get_event_dir, get_kdata_dir_csv, \
    get_tick_path_csv
from fooltrader.items import SecurityItem
from fooltrader.settings import STOCK_START_CODE, STOCK_END_CODE, TIME_FORMAT_DAY

logger = logging.getLogger(__name__)


def mkdir_for_security(item):
    fuquan_kdata_dir = get_kdata_dir(item, True)
    if not os.path.exists(fuquan_kdata_dir):
        os.makedirs(fuquan_kdata_dir)

    finance_dir = get_finance_dir(item)
    if not os.path.exists(finance_dir):
        os.makedirs(finance_dir)

    tick_dir = get_tick_dir(item)
    if not os.path.exists(tick_dir):
        os.makedirs(tick_dir)

    event_dir = get_event_dir(item)
    if not os.path.exists(event_dir):
        os.makedirs(event_dir)

    bfq_kdata_dir = get_kdata_dir_csv(item, 'bfq')
    if not os.path.exists(bfq_kdata_dir):
        os.makedirs(bfq_kdata_dir)

    hfq_kdata_dir = get_kdata_dir_csv(item, 'hfq')
    if not os.path.exists(hfq_kdata_dir):
        os.makedirs(hfq_kdata_dir)


def init_env():
    if not os.path.exists(settings.FILES_STORE):
        os.makedirs(settings.FILES_STORE)
    for item in get_security_items():
        mkdir_for_security(item)


def chrome_copy_header_to_dict(src):
    lines = src.split('\n')
    header = {}
    if lines:
        for line in lines:
            try:
                index = line.index(':')
                key = line[:index]
                value = line[index + 1:]
                if key and value:
                    header.setdefault(key.strip(), value.strip())
            except Exception:
                pass
    return header


def get_security_items(start=STOCK_START_CODE, end=STOCK_END_CODE):
    for item in itertools.chain(get_sh_security_item(),
                                get_sz_security_item()):
        if start <= item['code'] <= end:
            yield item


def generate_csv_line(*items):
    if items:
        result = items[0]
        for item in items[1:]:
            result = (result + ',' + item)
        return result
    return ''


def get_sz_security_item():
    path = get_sz_stock_list_path()
    wb = openpyxl.load_workbook(path)
    for name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(name)
        max_row, max_column = sheet.max_row, sheet.max_column
        for i in range(2, max_row):
            code = sheet.cell(row=i, column=1).value
            name = sheet.cell(row=i, column=2).value
            list_date = sheet.cell(row=i, column=8).value
            # ignore just in B
            if not list_date:
                continue
            yield SecurityItem(id=gen_security_id('stock', 'sz', code), type='stock', exchange='sz', code=code,
                               name=name, listDate=list_date)


def gen_security_id(type, exchange, code):
    return type + '_' + exchange + '_' + code;


def get_security_item(exchange):
    if exchange == 'sz':
        return get_sz_security_item()
    elif exchange == 'sh':
        return get_sh_security_item()


def get_sh_security_item():
    path = get_sh_stock_list_path()
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    with open(path, encoding=encoding) as fr:
        lines = fr.readlines()
        for line in lines[1:]:
            code, name, _, _, list_date, _, _ = line.split()
            yield SecurityItem(id=gen_security_id('stock', 'sh', code), type='stock', exchange='sh', code=code,
                               name=name, listDate=list_date)


def get_tick_items(security_item):
    for trading_date in get_trading_dates(security_item):
        tick_path = get_tick_path(security_item, trading_date)
        if os.path.exists(tick_path):
            yield get_tick_item(tick_path, trading_date, security_item)


def get_kdata_item_with_date(security_item, the_date_str):
    the_date = get_datetime(the_date_str)
    the_year_quarter = get_year_quarter(the_date)
    data_path = get_kdata_path(security_item, the_year_quarter[0], the_year_quarter[1], False)

    with open(data_path) as data_file:
        kdata_jsons = json.load(data_file)
        for kdata_json in kdata_jsons:
            if kdata_json['timestamp'] == the_date_str:
                return kdata_json


# 对于开盘涨停的，算作买盘tick
def kdata_to_tick(security_item, kdata_json):
    str = '''成交时间	成交价	价格变动	成交量(手)	成交额(元)	性质
{}	{}	--	{}	{}	{}'''.format('09:25:00', kdata_json['high'], int(kdata_json['volume']) / 100,
                                           kdata_json['turnover'], '买盘')
    return str


def get_kdata_items(security_item, houfuquan=False):
    dir = get_kdata_dir(security_item, houfuquan)
    if os.path.exists(dir):
        files = [os.path.join(dir, f) for f in os.listdir(dir) if
                 (f != "all_dayk.json" and os.path.isfile(os.path.join(dir, f)))]

        for f in sorted(files):
            with open(f) as data_file:
                kdata_jsons = json.load(data_file)
                for kdata_json in reversed(kdata_jsons):
                    yield kdata_json


def get_downloaded_tick_dates(security_item):
    dir = get_tick_dir(security_item)
    return [f[:f.index('.')] for f in os.listdir(dir) if os.path.isfile(os.path.join(dir, f))]


def get_tick_item(path, the_date, security_item):
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    with open(path, encoding=encoding) as fr:
        lines = fr.readlines()
        for line in reversed(lines[1:]):
            tmp_timestamp, price, tmp_change, volume, turnover, tmp_direction = line.split()
            # timestamp = datetime.datetime.strptime(the_date + tmp_timestamp, '%Y-%m-%d%H:%M:%S')
            timestamp = the_date + " " + tmp_timestamp
            change = 0
            if not tmp_change == '--':
                change = float(tmp_change)
            direction = 0
            if tmp_direction == '买盘':
                direction = 1
            elif tmp_direction == '卖盘':
                direction = -1

            # yield SecurityItem(code_id='sh' + code, code=code, name=name, list_date=list_date, exchange='sh',
            #                    type='stock')
            # yield generate_csv_line(code, name, list_date, 'sh', 'stock')
            yield {"securityId": security_item['id'],
                   "code": security_item['code'],
                   "timestamp": timestamp,
                   "price": price,
                   "change": change,
                   "direction": direction,
                   "volume": volume,
                   "turnover": turnover}


def detect_encoding(url):
    import urllib.request
    from chardet.universaldetector import UniversalDetector

    usock = urllib.request.urlopen(url)
    detector = UniversalDetector()
    for line in usock.readlines():
        detector.feed(line)
        if detector.done: break
    detector.close()
    usock.close()
    return detector.result.get('encoding')


def setup_env():
    pass


def merge_ths_kdata(security_item, dates):
    ths_kdata = {}
    ths_fuquan_kdata = {}

    try:
        with open(get_kdata_path_ths(security_item)) as data_file:
            ths_items = json.load(data_file)
            for item in ths_items:
                if item["timestamp"] in dates:
                    ths_kdata[item["timestamp"]] = item

        with open(get_kdata_path_ths(security_item, True)) as data_file:
            ths_items = json.load(data_file)
            for item in ths_items:
                if item["timestamp"] in dates:
                    ths_fuquan_kdata[item["timestamp"]] = item

        year_quarter_map_dates = {}
        for the_date in dates:
            year, quarter = get_year_quarter(get_datetime(the_date))
            year_quarter_map_dates.setdefault((year, quarter), [])
            year_quarter_map_dates.get((year, quarter)).append(the_date)

        for year, quarter in year_quarter_map_dates.keys():
            for fuquan in (False, True):
                data_path = get_kdata_path(security_item, year, quarter, fuquan)
                data_exist = os.path.isfile(data_path)
                if data_exist:
                    with open(data_path) as data_file:
                        k_items = json.load(data_file)
                        if fuquan:
                            for the_date in year_quarter_map_dates.get((year, quarter)):
                                k_items.append(ths_fuquan_kdata[the_date])
                        else:
                            for the_date in year_quarter_map_dates.get((year, quarter)):
                                k_items.append(ths_kdata[the_date])
                    k_items = sorted(k_items, key=lambda item: item["timestamp"], reverse=True)

                    with open(data_path, "w") as f:
                        json.dump(k_items, f)


    except Exception as e:
        logger.error(e)


def get_base_trading_dates(item):
    dates = []

    dates_path_sse = get_trading_dates_path_sse(item)
    dates_path_ths = get_trading_dates_path_ths(item)

    try:
        dates_sse = []
        dates_ths = []
        if dates_path_sse:
            with open(dates_path_sse) as data_file:
                dates_sse = json.load(data_file)
        if dates_path_ths:
            with open(dates_path_ths) as data_file:
                dates_ths = json.load(data_file)

        dates_tmp = set(dates_sse) & set(dates_ths)
        dates = list(dates_tmp)

    except Exception as e:
        logger.error(e)

    dates.sort()
    return dates


def get_trading_dates(security_item):
    dates = []
    dates_path_sina = get_trading_dates_path(security_item)
    try:
        with open(dates_path_sina) as data_file:
            dates = json.load(data_file)

    except Exception as e:
        logger.error(e)

    if not dates:
        # init_trading_dates(security_item)
        return get_trading_dates_path(security_item)

    return dates


def is_available_tick(path):
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    try:
        with open(path, encoding=encoding) as fr:
            line = fr.readline()
            return u'成交时间', u'成交价', u'价格变动', u'成交量(手)', u'成交额(元)', u'性质' == line.split()
    except Exception:
        return False


def get_datetime(str):
    return datetime.datetime.strptime(str, TIME_FORMAT_DAY)


def get_year_quarter(time):
    return time.year, ((time.month - 1) // 3) + 1


def get_quarters(start, end=datetime.date.today()):
    start_year_quarter = get_year_quarter(start)
    current_year_quarter = get_year_quarter(end)
    if current_year_quarter[0] == start_year_quarter[0]:
        return [(current_year_quarter[0], x) for x in range(start_year_quarter[1], current_year_quarter[1] + 1)]
    elif current_year_quarter[0] - start_year_quarter[0] == 1:
        return [(start_year_quarter[0], x) for x in range(start_year_quarter[1], 5)] + \
               [(current_year_quarter[0], x) for x in range(1, current_year_quarter[1] + 1)]
    elif current_year_quarter[0] - start_year_quarter[0] > 1:
        return [(start_year_quarter[0], x) for x in range(start_year_quarter[1], 5)] + \
               [(x, y) for x in range(start_year_quarter[0] + 1, current_year_quarter[0]) for y in range(1, 5)] + \
               [(current_year_quarter[0], x) for x in range(1, current_year_quarter[1] + 1)]
    else:
        raise Exception("wrong start time:{}".format(start));


def fill_doc_type(doc_type, json_object):
    for key in json_object:
        doc_type[key] = json_object[key]


def to_float(str):
    try:
        return float(str)
    except Exception as e:
        return None


def get_exchange(code):
    if code >= '333333':
        return 'sh'
    else:
        return 'sz'


def direction_to_int(direction):
    if direction == '买盘':
        return 1
    elif direction == '卖盘':
        return -1
    else:
        return 0


def sina_tick_to_csv(security_item, the_content, the_date):
    csv_path = get_tick_path_csv(security_item, the_date)
    df = pd.read_csv(the_content, sep='\s+', encoding='GB2312')
    df = df.loc[:, ['成交时间', '成交价', '成交量(手)', '成交额(元)', '性质']]
    df.columns = TICK_COLUNM
    df['direction'] = df['direction'].apply(lambda x: direction_to_int(x))
    df.to_csv(csv_path, index=False)
